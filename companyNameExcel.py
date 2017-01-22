#! python3
# Author simottar@gmail.com
# companyNameExcel.py - Prints the name for an icelandic registered company from the command line.



import json, requests, sys, openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# Compute file name from command line arguments.
if len(sys.argv) < 5:
    print('Usage: companyNameExcel.py FileName SheetName SsnColumn NameColumn')
    sys.exit()

# Get the file name
fileName = sys.argv[1]

# Get the sheet Name
sheetName = sys.argv[2]

# Get the column number in which we search for the name by SSN
columnSearchNumber = column_index_from_string(sys.argv[3])

# Get the column number in which we save the name
columnSaveNumber = column_index_from_string(sys.argv[4])



wb = openpyxl.load_workbook(fileName)
sheet = wb.get_sheet_by_name(sheetName)

for i in range(2, sheet.max_row):
    #Get the SSN from Excel
    ssn = sheet.cell(row=i, column=columnSearchNumber).value    
    print(ssn)
    # Download the JSON data from apis.is API.
    url ='http://apis.is/company?socialnumber=%s' % (ssn)
    response = requests.get(url)
    response.raise_for_status()

    # Load JSON data into a Python variable.
    companyData = json.loads(response.text)
    result = companyData['results']
    if result:
        name = result[0]['name']
        print(name)
        #update the desired column
        sheet.cell(row=i, column=columnSaveNumber, value= name) 

# Save the updated file
wb.save(fileName)